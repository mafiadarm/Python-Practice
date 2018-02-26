# -*- coding: utf-8 -*-
"""
==============================
   Date:           01_23_2018  9:29
   File Name:      /GitHub/updateProduce
   Creat From:     PyCharm
- - - - - - - - - - - - - - - 
   Description:    对excel的操作
==============================
"""

__author__ = 'Loffew'

import logging
import openpyxl

logging.basicConfig(level=logging.DEBUG, format=" %(asctime)s - %(levelname)s - %(message)s")  # [filename]


def pp_dbg(*args):
    return logging.debug(*args)

filePath = "C:\\Users\lo\Desktop\\10.0.0.175.xlsx"
wb = openpyxl.load_workbook(filePath)  # 打开文件  data_only = True 只显示值，隐藏公式[会有问题]
sheet = wb["24小时"]  # 获取工作薄

PRICE_UPDATES = {  # 需要更新的数据
    "访问网站": 777,  # 注意是否有 空格 等特殊符号
    "P2P流媒体 ": 888,
}

for i in range(2, sheet.max_row):
    produceName = sheet.cell(row=i, column=2).value  # row是行 column是列
    if produceName in PRICE_UPDATES:  # 如果键值在需要更新数据的字典里面
        sheet.cell(row=i, column=3).value = PRICE_UPDATES[produceName]  # 修改[column:row]后面一个空格的值
        pp_dbg((i, produceName, PRICE_UPDATES[produceName]))  # 标记修改，会后于所有print显示
# wb.save(filePath)

# 设置公式
wb["上班时间"]["A2"] = 200
wb["上班时间"]["B2"] = 300
wb["上班时间"]["C2"] = "=SUM(A2:B2)"
# wb.save(filePath)

# 执行以下步骤的时候，需要打开一下excel，注释掉前面的，单独执行下面的，就可以显示结果，否则会丢失公式
'''
filePath = "C:\\Users\lo\Desktop\\10.0.0.175.xlsx"
wb = openpyxl.load_workbook(filePath, data_only=True)  # 打开文件  data_only = True 只显示值，隐藏公式[会有问题]
print(wb["上班时间"]["C2"].value)
'''

# 高 宽
sheet.row_dimensions[2].height = 70  # 高
sheet.column_dimensions["B"].width = 20  # 宽
# wb.save(filePath)

# 合并 拆分单元格
sheet.merge_cells("A2:D3")
sheet["A2"] = "A2:D3 together"

sheet.unmerge_cells("A1:B1")
# wb.save(filePath)

# 冻结窗口
sheet.freeze_panes = "A2"
wb.save(filePath)

'''
from xlwt import Workbook
book = Workbook(encoding="utf-8")
sheet = book.add_sheet("provinces")
sheet.write(0,0,"1行A列")
sheet.write(0,1,"1行B列")
sheet.write(1,0,"2行A列")
sheet.write(1,1,"2行B列")
book.save("c:/Users/lo/Desktop/prvinces.xls")
'''
